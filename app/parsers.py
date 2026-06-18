"""Parse the two source xlsx files into rows the app can store."""
import re
from datetime import datetime
from openpyxl import load_workbook


COUNTRY_FROM_ORDER = {
    "UNITED KINGDOM": "UK",
    "CHANNEL ISLANDS": "UK",
    "IRELAND": "UK",
    "SWEDEN": "SWE",
    "NORWAY": "NOR",
}

PLAN_SHEETS = [
    ("DUK_26", "DUK", 2026),
    ("DUK_27", "DUK", 2027),
    ("UK_26", "UK", 2026),
    ("SWE_26", "SWE", 2026),
    ("NOR_26", "NOR", 2026),
    ("UK_27", "UK", 2027),
    ("SWE_27", "SWE", 2027),
    ("NOR_27", "NOR", 2027),
]


def _to_int(v):
    if v is None or v == "":
        return 0
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return 0


_JAN_RE = re.compile(r"^\s*JAN\s+(\d{4})\s*$", re.IGNORECASE)
_SECTION_END_LABELS = {"Dealer Inventory", "SHIPMENT", "Shipment"}


def _detect_plan_layout(ws):
    """Return (jan_col, year, model_col, super_col) by scanning the first ~6 header rows
    for a cell that reads 'JAN <year>'. Cols are 0-indexed for iter_rows tuples.
    """
    for r in range(1, 7):
        for c in range(1, min(ws.max_column + 1, 30)):
            v = ws.cell(r, c).value
            if not v:
                continue
            m = _JAN_RE.match(str(v))
            if m:
                jan_col_idx = c - 1
                return jan_col_idx, int(m.group(1)), jan_col_idx - 1, jan_col_idx - 2
    return None


def parse_plan(path):
    """Yield (country, plan_super, plan_model, year, month, qty) rows from the Sell In block
    of each country sheet. Layout is detected per sheet (UK_27 is shifted one column
    versus UK_26)."""
    wb = load_workbook(path, data_only=True)
    out = []
    for sheet_name, country, year in PLAN_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        layout = _detect_plan_layout(ws)
        if layout is None:
            continue
        jan_col, detected_year, model_col, super_col = layout
        if detected_year != year:
            year = detected_year  # trust the sheet over the sheet-name suffix

        in_block = False
        for row in ws.iter_rows(values_only=True):
            if len(row) <= jan_col + 11:
                continue
            label = row[model_col]
            label_str = str(label).strip() if label else ""
            if not in_block:
                if "Sell In" in label_str and str(year) in label_str:
                    in_block = True
                continue
            # In block.
            if label_str in _SECTION_END_LABELS:
                break
            model_h = row[model_col]
            if model_h in (None, "TOT", 0, "0"):
                continue
            super_g = row[super_col]
            model_clean = str(model_h).strip()
            super_clean = str(super_g).strip() if super_g else None
            if model_clean in ("TOT", "TOTAL BO") or model_clean.startswith("TOTAL"):
                continue
            for month_idx in range(12):
                qty = _to_int(row[jan_col + month_idx])
                if qty != 0:
                    out.append((country, super_clean, model_clean, year, month_idx + 1, qty))
    wb.close()
    return out


_MATERIAL_PREFIX_RE = re.compile(r"^\s*(\S+)")


def extract_material_prefix(material):
    if not material:
        return None
    m = _MATERIAL_PREFIX_RE.match(str(material))
    return m.group(1) if m else None


def _to_iso_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    try:
        return datetime.fromisoformat(str(v)).date().isoformat()
    except ValueError:
        return str(v)


_PERIOD_TEXT_RE = re.compile(
    r"(?i)(?:allocations\s+)?(\d{1,2})[.](\d{1,2})[.](\d{2,4})\s+(am|pm)",
)


def parse_period_text(text):
    """Parse 'DD.MM.YY am|pm' (with or without an 'Allocations ' prefix)
    into a sortable 'YYYY-MM-DD AM|PM' string. Returns None on miss.
    Used for both Allocations filenames and master-workbook sheet names."""
    if not text:
        return None
    m = _PERIOD_TEXT_RE.search(str(text))
    if not m:
        return None
    day, month, year_raw, period = m.groups()
    year = int(year_raw)
    if year < 100:
        year += 2000
    try:
        return f"{year:04d}-{int(month):02d}-{int(day):02d} {period.upper()}"
    except ValueError:
        return None


def parse_allocations(path):
    """Parse one or more allocation snapshots from a workbook.

    Returns (sheets, warnings):
      * sheets   — list of (sheet_name, report_date_or_None, rows) for every
                   sheet that yielded at least one row. For a single daily file
                   report_date is None (caller derives it); for a master
                   workbook it's parsed from the sheet name ('DD.MM.YY am|pm').
      * warnings — human-readable strings describing any sheet whose layout
                   differed from the standard export, so the caller can surface
                   them instead of silently dropping rows.

    Sheets in the master workbook are NOT all the same shape — columns get
    reordered, some have no header row, and some omit the 'Material' column
    entirely. Each sheet is parsed defensively (see _parse_allocations_sheet);
    anything lost or recovered-by-fallback is reported in `warnings`."""
    wb = load_workbook(path, data_only=True, read_only=True)
    out = []
    warnings = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows, diag = _parse_allocations_sheet(ws)
        # Loud diagnostics — only for sheets where something was off.
        if not diag["material_col"] and diag["skipped_no_material"]:
            warnings.append(
                f"'{sheet_name}': no 'Material' column — "
                f"{diag['skipped_no_material']} row(s) couldn't be mapped and were skipped."
            )
        elif diag["skipped_no_material"]:
            warnings.append(
                f"'{sheet_name}': {diag['skipped_no_material']} row(s) had no Material value and were skipped."
            )
        if rows and not diag["header"]:
            warnings.append(
                f"'{sheet_name}': no header row — read by column position ({len(rows)} rows recovered)."
            )
        if not rows:
            continue
        out.append((sheet_name, parse_period_text(sheet_name), rows))
    wb.close()
    return out, warnings


# A Material *description* looks like "HYMV2SP 26 EUR SP G-WT STD DMH" or
# "MS896ST 26 EUR MV G-..." — a code token, then a 2-digit model year, then the
# rest. Distinct from a Material *code* ("D50001526"), an order number, a PO
# number, a date, or a bike-model name. Used to locate the Material column on
# sheets that have no usable header.
_MATERIAL_DESC_RE = re.compile(r"^[A-Za-z0-9+]{2,}\s+2[0-9]\b")


def _norm_country(v):
    return COUNTRY_FROM_ORDER.get(str(v).strip().upper() if v else "", v)


def _parse_allocations_sheet(ws):
    """Parse one sheet of allocation rows. Returns (rows, diag).

    The master workbook's sheets aren't a fixed format, so we don't assume a
    rigid column order:
      * if a header row exists (a cell reading 'Order Number'), columns are
        mapped by NAME — robust to reordering and to missing columns;
      * if not, we fall back to the standard export's positional layout,
        anchored on the Material-description column detected by content.
    A row with no Material value can't be mapped to a plan model, so it's
    skipped and counted in diag rather than guessed at.

    diag = {header, material_col, rows, skipped_no_material} lets the caller
    warn loudly about non-standard sheets."""
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    diag = {"header": False, "material_col": False, "rows": 0, "skipped_no_material": 0}

    # Locate a header row: the first row containing an exact 'order number' cell.
    header_idx = None
    for i, row in enumerate(grid):
        if any(c is not None and str(c).strip().lower() == "order number" for c in row):
            header_idx = i
            break

    col = {}  # field -> column index
    if header_idx is not None:
        diag["header"] = True
        hdr = {}
        for j, c in enumerate(grid[header_idx]):
            if c is not None and str(c).strip():
                hdr.setdefault(str(c).strip().lower(), j)

        def find(*names):
            for n in names:
                if n in hdr:
                    return hdr[n]
            return None

        col = {
            "order_number": find("order number"),
            "material": find("material"),          # 'material code' won't match 'material'
            "bike_super_model": find("bike super model"),
            "bike_model": find("bike model"),
            "country": find("country", "sales country", "market"),
            "dealer_code": find("dealer code", "sold-to party"),
        }
        data = grid[header_idx + 1:]
    else:
        # Headerless: find the Material-description column by content, then read
        # neighbours using the standard layout's offsets from Material
        # (order=-2, super=+1, model=+2, country=+10, dealer_code=+12).
        data = [r for r in grid if any(c is not None and str(c).strip() for c in r)]
        counts = {}
        for r in data:
            for j, c in enumerate(r):
                if c is not None and _MATERIAL_DESC_RE.match(str(c).strip()):
                    counts[j] = counts.get(j, 0) + 1
        if counts:
            m = max(counts, key=lambda k: counts[k])
            col = {
                "order_number": m - 2 if m - 2 >= 0 else None,
                "material": m,
                "bike_super_model": m + 1,
                "bike_model": m + 2,
                "country": m + 10,
                "dealer_code": m + 12,
            }

    diag["material_col"] = col.get("material") is not None

    def cell(row, key):
        j = col.get(key)
        if j is None or j >= len(row):
            return None
        return row[j]

    out = []
    for row in data:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        material = cell(row, "material")
        order_raw = cell(row, "order_number")
        if not material and not order_raw:
            continue  # blank / spacer row
        if not material:
            diag["skipped_no_material"] += 1
            continue
        dealer_code = cell(row, "dealer_code")
        out.append({
            "order_number": str(order_raw) if order_raw else None,
            "material_prefix": extract_material_prefix(material),
            "material_full": str(material),
            "bike_super_model": cell(row, "bike_super_model"),
            "bike_model": cell(row, "bike_model"),
            "country": _norm_country(cell(row, "country")),
            "dealer_code": str(dealer_code) if dealer_code else None,
        })
        diag["rows"] += 1
    return out, diag


def parse_orders(path):
    """Yield order dicts from the Export sheet, filtered to open orders."""
    wb = load_workbook(path, data_only=True, read_only=True)
    if "Export" not in wb.sheetnames:
        wb.close()
        raise ValueError("Orders workbook has no 'Export' sheet")
    ws = wb["Export"]
    out = []
    header_seen = False
    for row in ws.iter_rows(values_only=True):
        if not header_seen:
            header_seen = True
            continue
        if not any(row):
            continue
        (order_number, _mat_code, material, super_model, bike_model,
         bike_color, bike_type, ec_status, _ec_date, request_date,
         _po_number, status_group, country_raw, dealer, dealer_code,
         _customer, order_create_date, confirmed_delivery_date) = (row + (None,) * 18)[:18]

        if status_group and str(status_group).strip().lower() != "open":
            continue

        out.append({
            "order_number": str(order_number) if order_number else None,
            "material_prefix": extract_material_prefix(material),
            "material_full": str(material) if material else None,
            "bike_super_model": super_model,
            "bike_model": bike_model,
            "bike_color": bike_color,
            "bike_type": bike_type,
            "end_customer_status": ec_status,
            "country": COUNTRY_FROM_ORDER.get(str(country_raw).strip().upper() if country_raw else "", country_raw),
            "dealer": dealer,
            "dealer_code": str(dealer_code) if dealer_code else None,
            "request_date": _to_iso_date(request_date),
            "order_creation_date": _to_iso_date(order_create_date),
            "confirmed_delivery_date": _to_iso_date(confirmed_delivery_date),
            "order_status_group": status_group,
        })
    wb.close()
    return out
